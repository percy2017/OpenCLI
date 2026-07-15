from pathlib import Path

from openpyxl import load_workbook

# Rows per emitted record. Keeps each record well under CHUNK_SIZE (512) when
# rows are short, while still letting long rows produce a single record that
# the chunker can split further.
_ROWS_PER_CHUNK = 25


def _row_to_text(row: tuple, header: tuple | None) -> str:
    cells = ["" if c is None else str(c) for c in row]
    if header is None:
        return "\t".join(cells)
    # Pair each value with its column name so retrieval matches on meaning,
    # not position. Empty column name means an unnamed trailing cell.
    pairs = []
    for col, val in zip(header, cells):
        col_name = col if col else ""
        pairs.append(f"{col_name}: {val}" if col_name else val)
    return " | ".join(pairs)


def _make_record(sheet_name: str, buffer: list[tuple], header: tuple | None) -> dict:
    lines = [_row_to_text(r, header) for r in buffer]
    return {
        "text": f"## Sheet: {sheet_name}\n" + "\n".join(lines),
        "sheet": sheet_name,
        "kind": "sheet_block",
    }


def load_xlsx(path: Path):
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            rows_iter = sheet.iter_rows(values_only=True)
            try:
                first = next(rows_iter)
            except StopIteration:
                continue

            header_cells = [("" if c is None else str(c).strip()) for c in first]
            has_header = any(h for h in header_cells)
            header: tuple | None = tuple(header_cells) if has_header else None

            # Buffer non-empty rows into blocks of _ROWS_PER_CHUNK so the
            # server-side splitter has a chance to chunk meaningfully instead
            # of receiving one 10k-line record per sheet.
            buffer: list[tuple] = []
            emitted = 0
            for row in rows_iter:
                cells = ["" if c is None else str(c) for c in row]
                if not any(c.strip() for c in cells):
                    continue
                buffer.append(tuple(cells))
                if len(buffer) >= _ROWS_PER_CHUNK:
                    yield _make_record(sheet_name, buffer, header)
                    emitted += len(buffer)
                    buffer = []
            if buffer:
                yield _make_record(sheet_name, buffer, header)
                emitted += len(buffer)

            # Fallback: if the sheet had no data rows but the first row was
            # a non-empty header, still yield the header so callers know the
            # sheet structure exists.
            if emitted == 0 and header is not None:
                yield {
                    "text": f"## Sheet: {sheet_name}\n" + " | ".join(header),
                    "sheet": sheet_name,
                    "kind": "sheet_header_only",
                }
    finally:
        wb.close()
